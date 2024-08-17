import pandas as pd
from decimal import Decimal
from openpyxl import Workbook

# Provided dictionary
data_all = {
    'FOL': {
        'SUB/L': '', 
        'SMT1': '', 
        'Mold1': '', 
        'SMT2': {
            'In': Decimal('10608'), 
            'Fail': Decimal('17'), 
            'Yield': '99.83%'
        }
    }, 
    'EOL': {
        'Mold2': '', 
        'SMT3': {
            'In': Decimal('4224'), 
            'Fail': Decimal('0'), 
            'Yield': '100.00%'
        }, 
        'LASER': {
            'In': Decimal('4224'), 
            'Fail': Decimal('0'), 
            'Yield': '100.00%'
        }, 
        'PKG Saw': {
            'In': Decimal('2816'), 
            'Fail': Decimal('0'), 
            'Yield': '100.00%'
        }, 
        'SPUTTER1': {
            'In': Decimal('13364'), 
            'Fail': Decimal('0'), 
            'Yield': '100.00%'
        }, 
        'SPUTTER2': {
            'In': Decimal('3508'), 
            'Fail': Decimal('0'), 
            'Yield': '100.00%'
        }, 
        'DMZ &FVI': {
            'In': Decimal('2464'), 
            'Fail': Decimal('0'), 
            'Yield': '100.00%'
        }
    }, 
    'TEST': {
        'SLT0': {
            'In': Decimal('20337'), 
            'Fail': Decimal('102'), 
            'Yield': '99.49%'
        }, 
        'SLT1': {
            'In': Decimal('14984'), 
            'Fail': Decimal('80'), 
            'Yield': '99.46%'
        }, 
        'SLT2': {
            'In': Decimal('11881'),
            'Fail': Decimal('41'), 
            'Yield': '99.65%'
        }, 
        'SLT3': '', 
        'AVI/TNR': {
            'In': Decimal('29720'), 
            'Fail': Decimal('62'), 
            'Yield': '99.79%', 
            'Hitter': [
                {'Hitter': {'Des': 'MISSING BALL', 'failQty': Decimal('1'), 'Rate': '0.00016%'}}, 
                {'Hitter': {'Des': ' MARKING REJECT', 'failQty': Decimal('4'), 'Rate': '0.00065%'}}, 
                {'Hitter': {'Des': 'CHIP OUT-TOP', 'failQty': Decimal('1'), 'Rate': '0.00016%'}}, 
                {'Hitter': {'Des': ' SCRATCH-TOP', 'failQty': Decimal('1'), 'Rate': '0.00016%'}}, 
                {'Hitter': {'Des': 'SPUTTER VOID/DENT', 'failQty': Decimal('1'), 'Rate': '0.00016%'}}, 
                {'Hitter': {'Des': 'PEEL OFF', 'failQty': Decimal('2'), 'Rate': '0.00032%'}}, 
                {'Hitter': {'Des': 'PEEL OFF', 'failQty': Decimal('43'), 'Rate': '0.00694%'}}, 
                {'Hitter': {'Des': 'MOLD FLASH', 'failQty': Decimal('4'), 'Rate': '0.00065%'}}, 
                {'Hitter': {'Des': 'BALL DAMAGE', 'failQty': Decimal('2'), 'Rate': '0.00032%'}}, 
                {'Hitter': {'Des': ' SCRATCH-BOTTOM', 'failQty': Decimal('1'), 'Rate': '0.00016%'}}
            ]
        }
    }
}

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Initialize current_row
current_row = 5

for group, stations in data_all.items():
    group_start_row = current_row
    ws.cell(row=current_row, column=1, value=group)
    current_row += 1
    for station, data in stations.items():
        station_start_row = current_row
        ws.cell(row=current_row, column=2, value=station)
        current_row += 1
        if data == "": 
            ws.cell(row=current_row-1, column=3, value=0)
            ws.cell(row=current_row-1, column=4, value=0)
            ws.cell(row=current_row-1, column=5, value="00.00%")
            ws.cell(row=current_row-1, column=7, value=0)
            ws.cell(row=current_row-1, column=8, value="00.00%")
            continue
        for info, detail in data.items():
            data_start_row = current_row
            if len(str(detail)) < 10:
                ws.cell(row=current_row, column=3, value=f"{info} : {detail}")
                current_row += 1
            else:
                ws.cell(row=current_row, column=3, value=f"{info} :")
                for hitter_no in detail:
                    detail_start_row = current_row
                    for hitter, hitter_info in hitter_no.items():
                        ws.cell(row=current_row, column=4, value=f'{hitter} :')
                        current_row += 1
                        hitter_column = 6
                        for key, value in hitter_info.items():
                            ws.cell(row=current_row, column=hitter_column, value=f'{value}')
                            hitter_column +=1
                current_row += 1

# Save the workbook
wb.save("testoutput.xlsx")
print("Exported")



