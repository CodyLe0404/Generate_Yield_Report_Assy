import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# Create a new workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Set the title of the worksheet
ws.title = "Yield & Hitters Review"

# Merge cells for the main title and set its value
ws.merge_cells('A2:D3')
main_title = ws['A2']
main_title.value = "Yield & Hitters Review"
main_title.alignment = Alignment(horizontal='left')
main_title.font = Font(bold=True)

ws.merge_cells('I1:J2')
second_main_title = ws['I1']
second_main_title.value = "M6 / Z6 / 050"
second_main_title.alignment = Alignment(horizontal='left')
second_main_title.font = Font(bold=True)

A5cell = ws['A5']
A5cell.value = "Yield Review Table"
A5cell.alignment = Alignment(horizontal='left')
A5cell.font = Font(bold=True)

ws.merge_cells('B5:E5')
B5cell_title = ws['B5']
B5cell_title.value = "639-18807"
B5cell_title.alignment = Alignment(horizontal='left')
B5cell_title.font = Font(bold=True)

ws.merge_cells('F5:J5')
B5cell_title = ws['F5']
B5cell_title.value = "24Hrs Top5 Hitters"
B5cell_title.alignment = Alignment(horizontal='left')
B5cell_title.font = Font(bold=True)

# Add column headers
headers = ['Dept./Group', 'Opr', 'Input # Qty', 'Fail Q\'ty', 'Engineering Yield', 'Hitters', 'Q\'ty', 'Defect Rate', 'Root Cause', 'Action']
ws.append(headers)

# Apply formatting to header row
for cell in ws[6]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# Set column widths (optional, adjust as needed)
column_widths = [15, 10, 12, 10, 18, 10, 10, 12, 12, 15]
for i, column_width in enumerate(column_widths, start=1):
    ws.column_dimensions[chr(64+i)].width = column_width

# Save the workbook to an Excel file
wb.save('formatted_excel.xlsx')
print("Exported")
