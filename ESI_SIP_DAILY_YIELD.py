import xlrd
import pyodbc
from openpyxl import Workbook
import pandas as pd
from xlutils.copy import copy
from xlwt import Workbook, XFStyle, Borders
from openpyxl.styles import Border, Side, Alignment, Font
from datetime import datetime, timedelta
import pytz
import base64
from bs4 import BeautifulSoup
import requests
import os
import glob
import configparser

timezone = pytz.timezone('Asia/Ho_Chi_Minh')
now = datetime.now(timezone)
yesterday = now - timedelta(days=1)
yesterday = str(yesterday.date())
today = str(now.date())
# yesterday = '2024-09-19'
# today = '2024-09-20'
Cur_Date = yesterday.replace("-","")

def get_data_group(cursor, device_no, Cur_Date, today):
    from_date = f'{Cur_Date}060000'
    to_date = f'{today.replace('-','')}055959'
    get_group_data = f"EXEC [GetGroupData_Assy] @DEVICE_TYPE_NO = '{device_no}', @FROM_DATE = '{from_date}', @TO_DATE = '{to_date}'"
    cursor.execute(get_group_data)
    group_data = cursor.fetchall()
    return group_data

def Get_AmkorID_SubID(cursor, DEVICE_TYPE_NO, CUR_DATE):
    get_hitter_data = f"EXEC [GetHitter_Assy] @DEVICE_TYPE_NO = '{DEVICE_TYPE_NO}', @CUR_DATE = '{CUR_DATE}'"
    cursor.execute(get_hitter_data)
    hitter_data = cursor.fetchall()
    return hitter_data

def data_24hrs_at_6am(list_hitter: list, yesterday: str, today: str): #Get data within 24 hours at 6 am every day
    # Define the start and end datetime range
    yesterday = yesterday.replace('-','/')
    today = today.replace('-','/')
    start_datetime = datetime.strptime(f'{yesterday} 06:00:00', '%Y/%m/%d %H:%M:%S')
    end_datetime = datetime.strptime(f'{today} 05:59:59', '%Y/%m/%d %H:%M:%S')

    # Filter the list based on the datetime range
    filtered_list = [desc for desc in list_hitter if start_datetime <= datetime.strptime(desc[4], '%Y/%m/%d %H:%M:%S') <= end_datetime]
    return filtered_list

def Get_Hitter(cursor, Device, Cur_Date, yesterday, today):
    list_hitter = []
    hitter_data = Get_AmkorID_SubID(cursor, Device, Cur_Date)
    Current_Date = yesterday.replace('-', '/')[:-1] 
    for row in hitter_data:
        amkorID, subID, cus_no, package = map(str, (row[0], row[1], row[-2], row[-1]))
        try: url = f'http://aav1ws01/eMES/sch/historyDefect.do?factoryID=80&siteID=1&wipAmkorID={amkorID}&wipAmkorSubID={subID}&pkg={package}&cust={cus_no}' #window
        except: url = f'http://10.201.16.21:9080//eMES/sch/historyDefect.do?factoryID=80&siteID=1&wipAmkorID={amkorID}&wipAmkorSubID={subID}&pkg={package}&cust={cus_no}' #linux
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')
        # Find the table by its tag
        table = soup.find('table')
        # Initialize a list to store the rows
        table_data = []

        # Iterate over each row in the table
        for row in table.find_all('tr'):
            # Get all the cells in the row
            cells = row.find_all(['td', 'th'])
            # Extract the text from each cell and store it in a list
            row_data = [cell.get_text(strip=True) for cell in cells]
            # Add the row data to the table data list
            table_data.append(row_data)

        # Print the extracted table data
        for row in table_data:
            data = [i for i in row if any(f'{Current_Date}' in cell for cell in row) and len(row) == 5]
            list_hitter.append(data)

    for i in list_hitter:
        if i != []:
            sql_get_group = f"SELECT Group_station FROM [MCSDB].[dbo].[Device_Data_Oper_Assy] WHERE Operation = '{i[1]}' and DeviceName = '{Device}'"
            cursor.execute(sql_get_group)
            group_data = cursor.fetchone()
            group_data = str(group_data)[2:-3]
            if group_data:
                i.append(group_data)

    list_hitter = [item for item in list_hitter if len(item) > 0]
    list_hitter_filtered = data_24hrs_at_6am(list_hitter, yesterday, today)

    return list_hitter_filtered
        
def Get_Yield(Yield):
    Yield = str(Yield).split('.')[0] + '.' + str(Yield).split('.')[1][:2] + '%'
    if int(Yield.split('.')[0]) >= 100:
        Yield = '100.00%'
        return Yield
    return Yield

# Define a dictionary to store the data
def generate_report_daily(cursor, device_no, Cur_Date, today):
    data_INACTIVE = get_data_group(cursor, device_no, Cur_Date, today)
    if data_INACTIVE == []:
        return 0
    data_dict = {}

    for index in data_INACTIVE:
        data_dict[index[0]] = {'In': index[1], 'Out': index[2], 'Yield': Get_Yield(index[3])}

    # Load the workbook and select the first sheet and define list of the keys
    if 'QM' not in device_no:   #For ESI
        rb = xlrd.open_workbook(r"C:\Workplace\Task\Support_Assy\Auto_Mail_Yield\sample_format\ESI_SIP_Sample_input.xls", formatting_info=True) #window
        # rb = xlrd.open_workbook(r"/home/testit/SRC/Source_2024/Support/ASSY_Generate_Yield_Report/sample_format/ESI_SIP_Sample_input.xls", formatting_info=True) #linux
        keys = ['SUB/L', 'SMT1', 'MOLD1', 'SMT2', 'MOLD2', 'SMT3', 'LASER', 'PKG Saw', 'SPUTTER1', 'SPUTTER2', 'DMZ &FVI', 'SLT0', 'SLT1', 'SLT2', 'SLT3', 'AVI/TNR']
    else:                       #For QORVO
        rb = xlrd.open_workbook(r"C:\Workplace\Task\Support_Assy\Auto_Mail_Yield\sample_format\QORVO_Sample_input.xls", formatting_info=True)
        # rb = xlrd.open_workbook(r"/home/testit/SRC/Source_2024/Support/ASSY_Generate_Yield_Report/sample_format/QORVO_Sample_input.xls", formatting_info=True) #window
        keys = ['2DSM', 'TOP SMT', 'TOP MOLD', 'BTM SMT', 'BTM MOLD', 'LASER', 'SMT Reball', 'PKG Saw', 'SPUTTER1', 'DMZ &FVI', 'SLT0', 'SLT1', 'SLT2', 'SLT3', 'AVI/TNR'] #linux

    #Get Limit Table
    sql_get_YLTB = f""" SELECT * FROM OPENQUERY([DATA400], 'SELECT YLYLIM FROM EMLIB.EMESTP04 
                WHERE YLPKG = ''M6'' AND YLDMS = ''Z6'' AND YLLEAD = ''050'' AND YLBUSN = ''A'' ')"""
    cursor.execute(sql_get_YLTB)
    data_ylmit = cursor.fetchone()
    yield_limit = str(data_ylmit[0])[:5] + '%'

    #WorkBook
    wb = copy(rb)
    sheet = wb.get_sheet(0)

    # Define the border style
    borders = Borders()
    borders.left = Borders.THIN
    borders.right = Borders.THIN
    borders.top = Borders.THIN
    borders.bottom = Borders.THIN
    style = XFStyle()
    style.borders = borders

    sheet.write(1, 0, '220', style)
    sheet.write(1, 1, device_no, style)
    if keys[0] == 'SUB/L':
        sheet.write(13, 2, yield_limit, style)
    else:
        sheet.write(12, 2, yield_limit, style)
    
    # Initialize the overall In and Out
    Overall_In = 0
    Overall_Out = 0

    # Modify cells from column D to H and row 4 to 17
    for row in range(3, len(keys) + 3):  # Rows 4 to 17 (0-indexed)
        key = keys[row - 3]
        if key in data_dict:
            sheet.write(row, 4, data_dict[key]['Yield'], style)
            sheet.write(row, 5, data_dict[key]['In'], style)
            sheet.write(row, 6, data_dict[key]['Out'], style)
            Overall_In += data_dict[key]['In']
            Overall_Out += data_dict[key]['Out']

    # Calculate the overall yield
    Overall_Yield = round((Overall_Out/Overall_In)*100,2)
    if int(Overall_Yield) >= 100:
        Overall_Yield = '100.00%'
    else:
        Overall_Yield = str(Overall_Yield) + '%'

    # Write the overall yield, in and out
    sheet.write(len(keys) + 4, 4, Overall_Yield, style)
    sheet.write(len(keys) + 4, 5, Overall_In, style)
    sheet.write(len(keys) + 4, 6, Overall_Out, style)
    fileName = f'{device_no}_{Cur_Date}_IO_DAILY_YIELD.xls'
    fileFolder = 'exported'     #window
    # fileFolder = '/home/testit/SRC/Source_2024/Support/ASSY_Generate_Yield_Report/exported'  #linux
    # Save the workbook
    wb.save(f'{fileFolder}/{fileName}')
    print(f"Exported -> {fileName}")
    return fileName

def generate_data_yield_summary(cursor, device_no, Cur_Date, yesterday, today):
    data_dict = {}
    data_dict_hitter = {}
    data_INACTIVE = get_data_group(cursor, device_no, Cur_Date, today)
    if data_INACTIVE == []:
        return 0
    
    for index in data_INACTIVE:
        failQty = index[1] - index[2]
        hit_type = index[0]
        hitter_info = {'In': int(index[1]), 'Fail' : int(failQty),'Yield': Get_Yield(index[3])}    
    # Append the hitter info to a list under the hit type key
        if hit_type not in data_INACTIVE:
            data_dict[hit_type] = hitter_info
        else:
            data_dict[hit_type].append(hitter_info)

    data_Hitter = Get_Hitter(cursor, device_no, Cur_Date, yesterday, today)
    flag = 0
    for ele in data_Hitter: # for index in data_Hitter:
        station = ele[-1]
        for index in data_INACTIVE:
            if index[0] == station:
                flag = 1
        if flag != 1:
            continue
        failDefect = int(ele[3])
        failStation = int(data_dict[station]['Fail'])
        if failStation == 0:
            rateDefect = '00.00%'
        else:
            rateDefect = round((failDefect/failStation)*100,2)
            rateDefect = str(rateDefect) + '%'
        dat_hitter = {
            'Hitter' : {
                'Des' : ele[2],
                'failQty' : int(ele[3]),
                'Rate' : rateDefect}}
        if station not in data_dict_hitter:
            data_dict_hitter[station] = []
            data_dict_hitter[station].append(dat_hitter)
            continue
        flag = 0
        len_data_station = len(data_dict_hitter[station])
        for ind in range(len_data_station):
            if ele[2] == data_dict_hitter[station][ind]['Hitter']['Des']:
                failQty_cumulative = data_dict_hitter[station][ind]['Hitter']['failQty'] + int(ele[3])
                data_dict_hitter[station][ind]['Hitter']['failQty'] = failQty_cumulative
                rateDefect = round((failQty_cumulative/failStation)*100,2)
                rateDefect = str(rateDefect) + '%'
                data_dict_hitter[station][ind]['Hitter']['Rate'] = rateDefect
                flag=1
        if flag != 1:
            data_dict_hitter[station].append(dat_hitter)

    if 'QM' not in device_no:
        stations = ['SUB/L','SMT1', 'Mold1', 'SMT2','Mold2', 'SMT3', 'LASER', 'PKG Saw', 'SPUTTER1', 'SPUTTER2', 'DMZ &FVI', 'SLT0', 'SLT1', 'SLT2', 'SLT3', 'AVI/TNR']
    else:
        stations = ['2DSM', 'TOP SMT', 'TOP MOLD', 'BTM SMT', 'BTM MOLD', 'LASER', 'SMT Reball', 'PKG Saw', 'SPUTTER1', 'DMZ &FVI', 'SLT0', 'SLT1', 'SLT2', 'SLT3', 'AVI/TNR']
    for station in stations:
        if station not in data_dict_hitter:
            continue
        data_dict[station]['Hitter'] = data_dict_hitter[station]
    return data_dict

def all_data_build(data, device_type):
    if 'QM' not in device_type:
        data_all = {
        'FOL': {key: "" for key in ['SUB/L', 'SMT1', 'MOLD1', 'SMT2']},
        'EOL': {key: "" for key in ['MOLD2', 'SMT3', 'LASER', 'PKG Saw', 'SPUTTER1', 'SPUTTER2', 'DMZ &FVI']},
        'TEST': {key: "" for key in ['SLT0', 'SLT1', 'SLT2', 'SLT3', 'AVI/TNR']}
        }
    else:
        data_all = {
        'FOL': {key: "" for key in ['2DSM', 'TOP SMT', 'TOP MOLD', 'BTM SMT']},
        'EOL': {key: "" for key in ['BTM MOLD', 'LASER', 'SMT Reball', 'PKG Saw', 'SPUTTER1', 'DMZ &FVI']},
        'TEST': {key: "" for key in ['SLT0', 'SLT1', 'SLT2', 'SLT3', 'AVI/TNR']}
        }

    for key in data:
        if key in data_all['FOL']:
            data_all['FOL'][key] = data[key]
        elif key in data_all['EOL']:
            data_all['EOL'][key] = data[key]
        elif key in data_all['TEST']:
            data_all['TEST'][key] = data[key]

    return data_all

def generate_yield_hitter_report(data_all, device_no, Cur_Date):
    # Tạo workbook và active worksheet
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # Set the title of the worksheet
    ws.title = "Yield & Hitters Review"

    # Merge cells for the main title and set its value
    ws.merge_cells('A2:D3')
    main_title = ws['A2']
    main_title.value = "Yield & Hitters Review"
    main_title.alignment = Alignment(horizontal='left')
    main_title.font = Font(bold=True)

    ws.merge_cells('I1:J2')
    second_main_title = ws['I1']
    second_main_title.value = f"M6 / Z6 / 050 - 220"
    second_main_title.alignment = Alignment(horizontal='left')
    second_main_title.font = Font(bold=True)

    A5cell = ws['A5']
    A5cell.value = "Yield Review Table"
    A5cell.alignment = Alignment(horizontal='left')
    A5cell.font = Font(bold=True)

    ws.merge_cells('B5:E5')
    B5cell_title = ws['B5']
    B5cell_title.value = f"{device_no} / {Cur_Date}"
    B5cell_title.alignment = Alignment(horizontal='left')
    B5cell_title.font = Font(bold=True)

    ws.merge_cells('F5:J5')
    B5cell_title = ws['F5']
    B5cell_title.value = "24Hrs Top5 Hitters"
    B5cell_title.alignment = Alignment(horizontal='left')
    B5cell_title.font = Font(bold=True)

    # Add column headers
    headers = ['Dept./Group', 'Opr', 'Input', 'Fail Q\'ty', 'Engineering Yield', 'Hitters', 'Q\'ty', 'Defect Rate', 'Root Cause', 'Action']
    ws.append(headers)

    # Apply formatting to header row
    for cell in ws[6]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Set column widths (optional, adjust as needed)
    column_widths = [18, 12, 10, 10, 18, 30, 10, 12, 12, 12]
    for i, column_width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64+i)].width = column_width

    # Initialize current_row
    current_row = 7
    for group, stations in data_all.items():
        group_start_row = current_row
        ws.cell(row=current_row, column=1, value=group)
        cell = ws.cell(row=current_row, column=1)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        for station, data in stations.items():
            station_start_row = current_row
            ws.cell(row=current_row, column=2, value=station)
            cell = ws.cell(row=current_row, column=2)
            cell.alignment = Alignment(vertical='center')
            if data == "":
                ws.cell(row=current_row, column=3, value=0)
                ws.cell(row=current_row, column=4, value=0)
                ws.cell(row=current_row, column=5, value="00.00%")
                ws.cell(row=current_row, column=7, value=0)
                ws.cell(row=current_row, column=8, value="00.00%")
                current_row += 1
                continue
            detail_column = 3
            for _, detail in data.items():
                data_start_row = current_row
                if len(str(detail)) < 10: 
                    ws.cell(row=current_row, column=detail_column, value=f"{detail}")
                    detail_column += 1
                    if detail_column == 6:
                        ws.cell(row=current_row, column=7, value=0)
                        ws.cell(row=current_row, column=8, value="00.00%")
                else:
                    for hitter_no in detail:
                        detail_start_row = current_row
                        for hitter, hitter_info in hitter_no.items():
                            hitter_column = 6
                            for key, value in hitter_info.items():
                                ws.cell(row=current_row, column=hitter_column, value=f'{value}')
                                hitter_column += 1
                        if hitter_no != detail[-1]:
                            current_row += 1
            
            ws.merge_cells(start_row=data_start_row, start_column=2, end_row=current_row, end_column=2)
            ws.merge_cells(start_row=data_start_row, start_column=3, end_row=current_row, end_column=3)
            ws.merge_cells(start_row=data_start_row, start_column=4, end_row=current_row, end_column=4)
            ws.merge_cells(start_row=data_start_row, start_column=5, end_row=current_row, end_column=5)
            current_row += 1
        ws.merge_cells(start_row=group_start_row, start_column=1, end_row=current_row-1, end_column=1)

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border    

    # Save the workbook
    fileName = f"{device_no}_{Cur_Date}_Yield_Hitter_Summary.xls"
    # fileFolder = '/home/testit/SRC/Source_2024/Support/ASSY_Generate_Yield_Report/exported'    #linux
    fileFolder = 'exported'                     #window
    wb.save(f"{fileFolder}/{fileName}")
    print(f"Exported -> {fileName}")
    return fileName

def convert_file_to_base64(filename):
    file_path = f"exported/{filename}"      #window
    # file_path = f'/home/testit/SRC/Source_2024/Support/ASSY_Generate_Yield_Report/exported/{filename}'    #linux
    with open(file_path, "rb") as file:
        encoded_string = base64.b64encode(file.read())
    return encoded_string.decode('utf-8')

def sending_email(list_attached):
    # toList = ['ATVPE@amkor.onmicrosoft.com']
    toList = ['Hiep.Letien@amkor.com']
    # toList = ['Khuong.Hoangminh@amkor.com','thuy.buithibich@amkor.com']
    # bccList = ['Hiep.Letien@amkor.com','Hoan.Nguyenvan@amkor.com']
    bccList = ['Hiep.Letien@amkor.com']
    dictionary_email = {
        "mailPriority": "NORMAL",
        "sender": "Assy.SummaryYield@amkor.com",
        "subject": f"{Cur_Date}_ATV_VB11000_YIELD DAILY REPORT",
        "body": f"<h1>Assy Summary Yield Report on {Cur_Date}</h1>",
        "toMailList": toList,
        "ccMailList": [""],
        "bccMailList": bccList,
        "attachmentList": list_attached
    }
    request_API(dictionary_email)

def request_API(payload):
    import requests
    import json
    headers = {'Content-Type': 'application/json'}
    # Send the files to the API
    response = requests.post("http://10.201.12.31:8004/Common/Send_Email", data=json.dumps(payload), headers=headers)
    print(response.text)

def delete_report_exported():
    folder_path = r'C:\Workplace\Task\Support_Assy\Auto_Mail_Yield\exported'    #window
    # folder_path = '/home/testit/SRC/Source_2024/Support/ASSY_Generate_Yield_Report/exported/'     #linux
    excel_files = glob.glob(os.path.join(folder_path, '*.xls'))
    for file in excel_files:
        os.remove(file)
    print(f"Deleted all")

def connect_database_window(host, port, user, password, database): #DRIVER={ODBC Driver 17 for SQL Server};SERVER=10.201.21.84,50150;DATABASE=ATV_Common;UID=cimitar2;PWD=TFAtest1!2!
    connectionStr = "DRIVER={ODBC Driver 17 for SQL Server}" 
    connectionStr += ";SERVER=" + host + "," + port 
    connectionStr +=  ";DATABASE=" + database 
    connectionStr += ";UID=" + user 
    connectionStr += ";PWD=" + password 
    return connectionStr

def connect_data_linux(host, port, user, password, database):
    #cnxn = pyodbc.connect("DRIVER=/opt/microsoft/msodbcsql18/lib64/libmsodbcsql-18.3.so.2.1;UID=cimitar2;PWD=TFAtest1!2!;Database=ATV_Common;Server=10.201.21.84,50150;TrustServerCertificate=yes;")
    connectionStr = "DRIVER=/opt/microsoft/msodbcsql18/lib64/libmsodbcsql-18.3.so.2.1" 
    connectionStr += ";UID=" + user
    connectionStr += ";PWD=" + password
    connectionStr +=  ";DATABASE=" + database 
    connectionStr += ";SERVER=" + host + "," + port 
    connectionStr += ";TrustServerCertificate=yes;"
    return connectionStr

def main():
    config = configparser.ConfigParser()
    config.read('config.ini')   #window
    # config.read('/home/testit/SRC/Source_2024/Support/ASSY_Generate_Yield_Report/config.ini')       #linux
    db_config = config['Database']
    host = db_config['Server']
    port = db_config['Port']
    user = db_config['User']
    password = db_config['Password']
    database = db_config['Database']

    connectionStr = connect_database_window(host, port, user, password, database)   #window
    # connectionStr = connect_data_linux(host, port, user, password, database)      #linux
    cnxn = pyodbc.connect(connectionStr)
    cursor = cnxn.cursor()

    device_no_list = ['639-18807', '639-18808', 'QM76300', 'QM76309', 'QM76095']
    device_no_1 = ['QM76309']
    list_attached = []
    for device_no in device_no_list:
        report_daily = generate_report_daily(cursor, device_no, Cur_Date, today)
        if report_daily != 0:
            data = generate_data_yield_summary(cursor, device_no, Cur_Date, yesterday, today)
            all_data = all_data_build(data, device_no)
            yield_hitter_report = generate_yield_hitter_report(all_data, device_no, Cur_Date)
            base64_report_daily = convert_file_to_base64(report_daily)
            base64_hitter_report = convert_file_to_base64(yield_hitter_report)
            list_attached.append({
                "base64File": base64_report_daily,
                "fileName": report_daily,
                "mimeType" : "application/vnd.ms-excel"
            })
            list_attached.append({
                "base64File": base64_hitter_report,
                "fileName": yield_hitter_report,
                "mimeType" : "application/vnd.ms-excel"
            })
        else:
            print(f"Cannot generate report because {device_no} has no data on {Cur_Date}")
    cnxn.close()
    sending_email(list_attached)
    delete_report_exported()
if __name__ == '__main__':
    main()

